import csv
import math
from pathlib import Path

from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex, QVariant, QMimeType
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableView, QComboBox

from omniviewer.viewers.base import BaseViewer

## @brief Ленивая модель таблицы для QTableView.
class LazyTableModel(QAbstractTableModel):
    def __init__(self, reader_iterator, headers=None):
        super().__init__()
        self._data = []
        self._headers = headers or []
        self._iterator = reader_iterator
        self._exhausted = False
        self.fetchMore(QModelIndex())

    def rowCount(self, parent=QModelIndex()):
        if parent.isValid(): return 0
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        if parent.isValid(): return 0
        return len(self._headers)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid(): return QVariant()
        if role == Qt.ItemDataRole.DisplayRole:
            row, col = index.row(), index.column()
            if row < len(self._data):
                row_data = self._data[row]
                if col < len(row_data):
                    val = row_data[col]
                    return str(val) if val is not None else ""
        return QVariant()

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            if section < len(self._headers):
                return str(self._headers[section])
            return str(section)
        return super().headerData(section, orientation, role)

    def canFetchMore(self, parent=QModelIndex()):
        if parent.isValid(): return False
        return not self._exhausted

    def fetchMore(self, parent=QModelIndex()):
        if parent.isValid(): return
        batch = 500
        start = len(self._data)
        new_rows = []
        try:
            for _ in range(batch):
                row = next(self._iterator)
                new_rows.append(row)
        except StopIteration:
            self._exhausted = True
        
        if new_rows:
            # Check column count
            max_col = max(len(r) for r in new_rows)
            if max_col > len(self._headers):
                self.beginInsertColumns(QModelIndex(), len(self._headers), max_col - 1)
                self._headers.extend([f"Col {i}" for i in range(len(self._headers), max_col)])
                self.endInsertColumns()

            self.beginInsertRows(QModelIndex(), start, start + len(new_rows) - 1)
            self._data.extend(new_rows)
            self.endInsertRows()

## @brief Просмотрщик табличных данных.
class SpreadsheetViewer(BaseViewer):
    priority = 100

    def __init__(self):
        super().__init__()
        self._layout = QVBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        
        self._sheet_selector = QComboBox()
        self._sheet_selector.currentIndexChanged.connect(self._on_sheet_changed)
        self._sheet_selector.hide()
        self._layout.addWidget(self._sheet_selector)
        
        self._table_view = QTableView()
        self._layout.addWidget(self._table_view)
        
        self._path = None
        self._file_type = None
        self._workbook = None
        self._file_obj = None

    @classmethod
    def can_handle(cls, path: Path, mime_type: QMimeType) -> bool:
        name = mime_type.name()
        if name in [
            "text/csv",
            "text/tab-separated-values",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "application/vnd.oasis.opendocument.spreadsheet"
        ]:
            return True
        if path.suffix.lower() in [".csv", ".tsv", ".xlsx", ".xlsm", ".xls", ".ods"]:
            return True
        return False

    def load(self, path: Path) -> None:
        self._path = path
        suffix = path.suffix.lower()
        
        if suffix in [".csv", ".tsv"]:
            self._load_csv(path, suffix)
        elif suffix in [".xlsx", ".xlsm"]:
            self._load_xlsx(path)
        elif suffix == ".xls":
            self._load_xls(path)
        elif suffix == ".ods":
            self._load_ods(path)
        else:
            # Try to guess based on MIME type if no suffix
            if self.mime_type and ("csv" in self.mime_type.name() or "tab" in self.mime_type.name()):
                self._load_csv(path, ".csv")
            else:
                raise ValueError(f"Unsupported spreadsheet format: {suffix}")

    def _load_csv(self, path: Path, suffix: str):
        self._file_obj = open(path, "r", encoding="utf-8", errors="replace")
        
        # Sniff delimiter
        sample = self._file_obj.read(1024)
        self._file_obj.seek(0)
        
        delimiter = ","
        if suffix == ".tsv":
            delimiter = "\t"
        elif sample:
            try:
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except csv.Error:
                pass
                
        reader = csv.reader(self._file_obj, delimiter=delimiter)
        try:
            headers = next(reader)
        except StopIteration:
            headers = []
            
        model = LazyTableModel(reader, headers)
        self._table_view.setModel(model)

    def _load_xlsx(self, path: Path):
        import openpyxl
        self._workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self._setup_sheets(self._workbook.sheetnames)

    def _load_xls(self, path: Path):
        import xlrd
        self._workbook = xlrd.open_workbook(path)
        self._setup_sheets(self._workbook.sheet_names())

    def _load_ods(self, path: Path):
        import odf.opendocument
        import odf.table
        import odf.text
        
        self._workbook = odf.opendocument.load(path)
        sheets = self._workbook.spreadsheet.getElementsByType(odf.table.Table)
        sheet_names = [s.getAttribute("name") for s in sheets]
        self._setup_sheets(sheet_names)

    def _setup_sheets(self, names):
        if not names:
            return
            
        self._sheet_selector.blockSignals(True)
        self._sheet_selector.clear()
        self._sheet_selector.addItems(names)
        self._sheet_selector.blockSignals(False)
        
        if len(names) > 1:
            self._sheet_selector.show()
        else:
            self._sheet_selector.hide()
            
        self._on_sheet_changed(0)

    def _on_sheet_changed(self, index: int):
        if index < 0 or not self._workbook:
            return
            
        suffix = self._path.suffix.lower()
        if suffix in [".xlsx", ".xlsm"]:
            sheet_name = self._sheet_selector.itemText(index)
            sheet = self._workbook[sheet_name]
            iterator = sheet.iter_rows(values_only=True)
        elif suffix == ".xls":
            sheet = self._workbook.sheet_by_index(index)
            def gen():
                for rx in range(sheet.nrows):
                    yield sheet.row_values(rx)
            iterator = gen()
        elif suffix == ".ods":
            import odf.table
            import odf.text
            sheets = self._workbook.spreadsheet.getElementsByType(odf.table.Table)
            sheet = sheets[index]
            
            def get_text(cell):
                text_nodes = cell.getElementsByType(odf.text.P)
                return " ".join(str(p) for p in text_nodes)
                
            def gen():
                rows = sheet.getElementsByType(odf.table.TableRow)
                for row in rows:
                    cells = row.getElementsByType(odf.table.TableCell)
                    yield [get_text(c) for c in cells]
            iterator = gen()
        else:
            return
            
        try:
            headers = next(iterator)
        except StopIteration:
            headers = []
            
        model = LazyTableModel(iterator, headers)
        self._table_view.setModel(model)
